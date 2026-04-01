"""
=============================================================================
  Analyse Hydro-Québec — Données 15 minutes
  Contexte : Audit énergétique client / Outil interne Clauger
=============================================================================
"""

import re
import io
import unicodedata

import numpy as np
import pandas as pd
import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
import matplotlib.ticker as mticker
import matplotlib.patches as mpatches
import seaborn as sns
import streamlit as st

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.drawing.image import Image as XLImage
from openpyxl.utils import get_column_letter

# ─────────────────────────────────────────────────────────────────────────────
# CONFIG PAGE
# ─────────────────────────────────────────────────────────────────────────────
st.set_page_config(
    page_title="Analyse HQ — 15 min",
    page_icon="⚡",
    layout="wide",
)

# Palette visuelle globale (cohérente dans tous les graphiques)
PALETTE = {
    "primaire":  "#1F4E79",
    "accent":    "#2196F3",
    "vert":      "#388E3C",
    "orange":    "#F57C00",
    "rouge":     "#C62828",
    "gris":      "#90A4AE",
    "hiver":     "#1565C0",
    "automne":   "#E65100",
    "printemps": "#2E7D32",
    "ete":       "#F9A825",
}

plt.rcParams.update({
    "font.family":   "DejaVu Sans",
    "axes.spines.top":   False,
    "axes.spines.right": False,
    "axes.grid":         True,
    "grid.alpha":        0.25,
    "grid.linestyle":    "--",
})

# ─────────────────────────────────────────────────────────────────────────────
# STYLE EXCEL (identique rapport capteurs — bleu navy)
# ─────────────────────────────────────────────────────────────────────────────
XL_TITRE_BG   = "0D1F3C"
XL_SOUS_BG    = "2E4057"
XL_SECTION_BG = "1F3864"
XL_HEADER_BG  = "1F4E79"
XL_WHITE      = "FFFFFF"
XL_ROW_EVEN   = "F2F2F2"
XL_ROW_ODD    = "FFFFFF"
XL_VERT_BG    = "C6EFCE";  XL_VERT_FG  = "006100"
XL_ORANGE_BG  = "FFC000";  XL_ORANGE_FG= "7F4800"
XL_ROUGE_BG   = "FFC7CE";  XL_ROUGE_FG = "9C0006"
XL_FONT       = "Arial"


def _thin():
    s = Side(style="thin", color="BFBFBF")
    return Border(left=s, right=s, top=s, bottom=s)

def _med_bottom():
    t = Side(style="thin",   color="BFBFBF")
    m = Side(style="medium", color=XL_HEADER_BG)
    return Border(left=t, right=t, top=t, bottom=m)

def xl_titre(ws, row, text, ncols, bg=XL_TITRE_BG, size=14):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=ncols)
    c = ws.cell(row=row, column=1, value=text)
    c.font = Font(name=XL_FONT, bold=True, size=size, color=XL_WHITE)
    c.fill = PatternFill("solid", fgColor=bg)
    c.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[row].height = 22

def xl_sous_titre(ws, row, text, ncols):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=ncols)
    c = ws.cell(row=row, column=1, value=text)
    c.font = Font(name=XL_FONT, size=10, color=XL_WHITE)
    c.fill = PatternFill("solid", fgColor=XL_SOUS_BG)
    c.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[row].height = 16

def xl_section(ws, row, text, ncols):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=ncols)
    c = ws.cell(row=row, column=1, value=text)
    c.font = Font(name=XL_FONT, bold=True, size=11, color=XL_WHITE)
    c.fill = PatternFill("solid", fgColor=XL_SECTION_BG)
    c.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[row].height = 18

def xl_headers(ws, row, headers):
    for ci, h in enumerate(headers, 1):
        c = ws.cell(row=row, column=ci, value=h)
        c.font = Font(name=XL_FONT, bold=True, size=10, color=XL_WHITE)
        c.fill = PatternFill("solid", fgColor=XL_HEADER_BG)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = _med_bottom()
    ws.row_dimensions[row].height = 30

def xl_data_row(ws, row_idx, data, statut_ci=None):
    bg = XL_ROW_EVEN if row_idx % 2 == 0 else XL_ROW_ODD
    for ci, val in enumerate(data, 1):
        c = ws.cell(row=row_idx, column=ci, value=val)
        c.font = Font(name=XL_FONT, size=10)
        c.border = _thin()
        if statut_ci and ci == statut_ci:
            if val == "🟢":
                c.fill = PatternFill("solid", fgColor=XL_VERT_BG)
                c.font = Font(name=XL_FONT, size=10, color=XL_VERT_FG, bold=True)
            elif val == "🟠":
                c.fill = PatternFill("solid", fgColor=XL_ORANGE_BG)
                c.font = Font(name=XL_FONT, size=10, color=XL_ORANGE_FG, bold=True)
            elif val == "🔴":
                c.fill = PatternFill("solid", fgColor=XL_ROUGE_BG)
                c.font = Font(name=XL_FONT, size=10, color=XL_ROUGE_FG, bold=True)
            else:
                c.fill = PatternFill("solid", fgColor=bg)
        else:
            c.fill = PatternFill("solid", fgColor=bg)
        c.alignment = Alignment(
            horizontal="left" if ci == 1 else "center",
            vertical="center", wrap_text=False
        )

def xl_auto_width(ws, ncols, df):
    for ci in range(1, ncols + 1):
        col_name = df.columns[ci - 1]
        vals = df.iloc[:, ci - 1].astype(str).tolist()
        max_len = max(len(str(col_name)), max((len(v) for v in vals), default=0))
        ws.column_dimensions[get_column_letter(ci)].width = min(max(max_len + 2, 10), 55)

def xl_write_df(ws, titre, sous_titre, section, df, statut_col=None, start_row=1):
    ncols = len(df.columns)
    r = start_row
    xl_titre(ws, r, titre, ncols);          r += 1
    xl_sous_titre(ws, r, sous_titre, ncols); r += 1
    r += 1  # vide
    if section:
        xl_section(ws, r, section, ncols);  r += 1
    xl_headers(ws, r, list(df.columns));    r += 1
    header_row = r - 1
    s_ci = (df.columns.get_loc(statut_col) + 1) if statut_col and statut_col in df.columns else None
    for _, row_data in df.iterrows():
        xl_data_row(ws, r, list(row_data), statut_ci=s_ci)
        r += 1
    xl_auto_width(ws, ncols, df)
    ws.freeze_panes = ws.cell(row=header_row + 1, column=1)
    return r


# ─────────────────────────────────────────────────────────────────────────────
# HELPERS — LECTURE / NETTOYAGE
# ─────────────────────────────────────────────────────────────────────────────
def norm_col(s: str) -> str:
    s = str(s).strip()
    s = unicodedata.normalize("NFKD", s).encode("ascii", "ignore").decode("ascii")
    return re.sub(r"\s+", " ", s).lower()

def read_file(uf) -> pd.DataFrame:
    name = uf.name.lower()
    if name.endswith(".csv"):
        for enc in ["ISO-8859-1", "utf-8", "cp1252"]:
            try:
                uf.seek(0)
                return pd.read_csv(uf, encoding=enc, sep=";")
            except Exception:
                pass
        uf.seek(0)
        return pd.read_csv(uf, encoding="ISO-8859-1", sep=",")
    uf.seek(0)
    return pd.read_excel(uf)

def detect_columns(df: pd.DataFrame):
    cols = list(df.columns)
    date_keywords  = ["date et heure", "horodate", "timestamp", "heure", "periode", "date"]
    power_keywords = ["puissance", "power", "appel", "demande", "charge"]

    date_col = next(
        (c for c in cols if any(k in norm_col(c) for k in ["date et heure", "horodate", "timestamp"])),
        next((c for c in cols if any(k in norm_col(c) for k in date_keywords)), None)
    )

    power_col = None
    kw_candidates = [c for c in cols if "kw" in norm_col(c) and any(k in norm_col(c) for k in power_keywords)]
    # priorité aux colonnes "réelle / mesurée"
    power_col = next(
        (c for c in kw_candidates if any(k in norm_col(c) for k in ["reel", "reelle", "mesuree", "mesur"])),
        kw_candidates[0] if kw_candidates else None
    )
    if power_col is None:
        power_col = next((c for c in cols if "kw" in norm_col(c)), None)

    return date_col, power_col

def clean_df(df: pd.DataFrame, date_col: str, power_col: str) -> pd.DataFrame:
    d = df[[date_col, power_col]].copy()
    d.rename(columns={date_col: "ts", power_col: "kW"}, inplace=True)
    d["kW"] = (
        d["kW"].astype(str)
        .str.replace(" ", "", regex=False)
        .str.replace(",", ".", regex=False)
    )
    d["kW"] = pd.to_numeric(d["kW"], errors="coerce")
    d["ts"] = pd.to_datetime(d["ts"], errors="coerce")
    d = d.dropna().drop_duplicates(subset=["ts"]).sort_values("ts").set_index("ts")
    return d

# ─────────────────────────────────────────────────────────────────────────────
# HELPERS — CALCULS
# ─────────────────────────────────────────────────────────────────────────────
def add_kwh(df: pd.DataFrame) -> pd.DataFrame:
    d = df.sort_index().copy()
    dh = d.index.to_series().diff().dt.total_seconds().div(3600).fillna(0).clip(lower=0)
    d["delta_h"] = dh
    d["kWh"]     = d["kW"] * dh
    return d

def compute_palier(pmax: float) -> int:
    if pmax <= 500:  return 500
    if pmax <= 700:  return 700
    if pmax <= 1000: return 1000
    return (int(pmax // 100) + 1) * 100

def enrich(df: pd.DataFrame) -> tuple:
    d = df.copy()
    pmax   = float(d["kW"].max())
    palier = compute_palier(pmax)
    d["Palier_kW"]  = palier
    d["Ecart_kW"]   = (palier - d["kW"]).clip(lower=0)
    d["FU_%"]       = d["kW"] / palier * 100

    # Saison (Québec)
    def saison(m):
        if m in [12, 1, 2]:  return "Hiver"
        if m in [3, 4, 5]:   return "Printemps"
        if m in [6, 7, 8]:   return "Été"
        return "Automne"

    d["Mois"]   = d.index.month
    d["Annee"]  = d.index.year
    d["Heure"]  = d.index.hour
    d["JourSemaine"] = d.index.dayofweek        # 0=Lun … 6=Dim
    d["TypeJour"]    = d["JourSemaine"].apply(lambda x: "Semaine" if x < 5 else "Week-end")
    d["Saison"]      = d["Mois"].apply(saison)
    d["Date"]        = d.index.date
    return d, palier

def median_step_min(idx: pd.DatetimeIndex) -> float:
    deltas = idx.to_series().diff().dropna()
    return float(deltas.median().total_seconds() / 60) if not deltas.empty else float("nan")

def fu_pondere(df: pd.DataFrame) -> float:
    den = df["delta_h"].sum()
    return float((df["FU_%"] * df["delta_h"]).sum() / den) if den > 0 else float("nan")

# ─────────────────────────────────────────────────────────────────────────────
# HELPERS — SAUVEGARDE FIGURES
# ─────────────────────────────────────────────────────────────────────────────
fig_buffers: dict = {}

def savefig(fig, key: str, dpi: int = 160):
    buf = io.BytesIO()
    fig.savefig(buf, format="png", dpi=dpi, bbox_inches="tight")
    buf.seek(0)
    fig_buffers[key] = buf
    plt.close(fig)

# ─────────────────────────────────────────────────────────────────────────────
# SIDEBAR + UPLOAD
# ─────────────────────────────────────────────────────────────────────────────
st.title("⚡ Analyse Hydro-Québec — Données 15 minutes")

with st.sidebar:
    st.header("⚙️ Paramètres")
    seuil_anomalie = st.slider(
        "Seuil anomalie (× écart-type au-dessus de la moyenne)", 2.0, 5.0, 3.0, 0.5,
        help="Un point est signalé comme pic si sa valeur dépasse : moyenne + N × σ"
    )
    bin_rep = st.selectbox("Largeur de tranche (répartition puissances)", [10, 25, 50, 100], index=1)
    st.divider()
    st.caption("Contexte : Audit énergétique / Clauger")

uploaded_files = st.file_uploader(
    "📂 Importe tes fichiers Hydro-Québec (CSV ou XLSX) — plusieurs fichiers acceptés",
    type=["csv", "xlsx"],
    accept_multiple_files=True
)

if not uploaded_files:
    st.info("⬆️ Importe au moins 1 fichier pour démarrer l'analyse.")
    st.stop()

# ─────────────────────────────────────────────────────────────────────────────
# LECTURE + NETTOYAGE
# ─────────────────────────────────────────────────────────────────────────────
cleaned_list = []
rejected     = []

for uf in uploaded_files:
    try:
        df_raw = read_file(uf)
        st.caption(f"📄 `{uf.name}` — colonnes : {list(df_raw.columns)}")

        date_col, power_col = detect_columns(df_raw)

        if date_col is None or power_col is None:
            st.warning(f"⚠️ Détection auto impossible pour `{uf.name}`. Sélection manuelle :")
            cols_list = list(df_raw.columns)
            date_col  = st.selectbox(f"Colonne DATE — {uf.name}",  cols_list, key=f"d_{uf.name}")
            power_col = st.selectbox(f"Colonne kW  — {uf.name}",  cols_list, key=f"p_{uf.name}")

        df_c = clean_df(df_raw, date_col, power_col)
        if df_c.empty:
            rejected.append((uf.name, "DataFrame vide après nettoyage"))
            continue

        df_c["_src"] = uf.name
        cleaned_list.append(df_c)

        with st.expander(f"Aperçu `{uf.name}` ({len(df_c):,} lignes)"):
            st.dataframe(df_c.reset_index().head(30), use_container_width=True)

    except Exception as e:
        rejected.append((uf.name, str(e)))

if rejected:
    with st.expander("⚠️ Fichiers rejetés"):
        for n, r in rejected:
            st.write(f"- **{n}** → {r}")

if not cleaned_list:
    st.error("⛔ Aucun fichier valide après nettoyage.")
    st.stop()

# ─────────────────────────────────────────────────────────────────────────────
# CONSTRUCTION DU DATAFRAME FINAL
# ─────────────────────────────────────────────────────────────────────────────
df = pd.concat(cleaned_list).sort_index()
df = df[~df.index.duplicated(keep="first")]
df = add_kwh(df)
df, palier = enrich(df)

step_min  = median_step_min(df.index)
pmax      = float(df["kW"].max())
e_kwh     = float(df["kWh"].sum())
h_total   = float(df["delta_h"].sum())
fu        = fu_pondere(df)
periode   = f"{df.index.min().strftime('%d %b %Y')} – {df.index.max().strftime('%d %b %Y')}"

# ─────────────────────────────────────────────────────────────────────────────
# SECTION 1 : KPI GLOBAUX
# ─────────────────────────────────────────────────────────────────────────────
st.divider()
st.subheader("📊 Indicateurs globaux")

k1, k2, k3, k4, k5, k6 = st.columns(6)
k1.metric("Appel de pointe",    f"{pmax:,.1f} kW")
k2.metric("Palier contractuel", f"{palier:,.0f} kW")
k3.metric("Énergie totale",     f"{e_kwh/1000:,.1f} MWh")
k4.metric("Heures couvertes",   f"{h_total:,.0f} h")
k5.metric("Facteur d'utilisation", f"{fu:.1f} %")
k6.metric("Pas médian",         f"{step_min:.1f} min")

st.caption(f"📅 Période : {periode}  |  Lignes : {len(df):,}  |  Source(s) : {', '.join(df['_src'].unique())}")

# ─────────────────────────────────────────────────────────────────────────────
# SECTION 2 : CONSOMMATION MENSUELLE & ANNUELLE
# ─────────────────────────────────────────────────────────────────────────────
st.divider()
st.subheader("📅 Consommation mensuelle et annuelle")

mon = pd.DataFrame({
    "Énergie_MWh":    df["kWh"].resample("ME").sum() / 1000,
    "P_max_kW":       df["kW"].resample("ME").max(),
    "P_moy_kW":       df["kW"].resample("ME").mean(),
    "FU_moy_%":       df["FU_%"].resample("ME").mean(),
    "Heures_couv":    df["delta_h"].resample("ME").sum(),
}).dropna()

# FU mensuel (formule Hydro-Québec : énergie / (P_max × h_mois))
if not mon.empty:
    mon["FU_HQ_%"] = (mon["Énergie_MWh"] * 1000) / (mon["P_max_kW"] * mon["Heures_couv"]) * 100
    mon["Mois_label"] = mon.index.strftime("%b %Y")

col_g1, col_g2 = st.columns(2)

with col_g1:
    if not mon.empty:
        fig, ax = plt.subplots(figsize=(9, 4))
        bars = ax.bar(mon["Mois_label"], mon["Énergie_MWh"],
                      color=PALETTE["primaire"], alpha=0.85, zorder=3)
        ax.set_title("Consommation mensuelle (MWh)", fontsize=12, fontweight="bold")
        ax.set_ylabel("MWh")
        ax.tick_params(axis="x", rotation=45)
        # Annotation valeurs
        for bar, val in zip(bars, mon["Énergie_MWh"]):
            ax.text(bar.get_x() + bar.get_width() / 2, bar.get_height() + 0.5,
                    f"{val:,.0f}", ha="center", va="bottom", fontsize=7, color="#333")
        plt.tight_layout()
        st.pyplot(fig)
        savefig(fig, "01_Consommation_mensuelle")

with col_g2:
    if not mon.empty:
        fig, ax = plt.subplots(figsize=(9, 4))
        ax.plot(mon["Mois_label"], mon["P_max_kW"],  marker="o", color=PALETTE["rouge"],
                label="P max (facturée)", linewidth=2)
        ax.plot(mon["Mois_label"], mon["P_moy_kW"],  marker="s", color=PALETTE["accent"],
                label="P moyenne (consommée)", linewidth=2)
        ax.axhline(palier, color=PALETTE["orange"], linestyle="--", linewidth=1.5,
                   label=f"Palier {palier} kW", alpha=0.8)
        ax.set_title("Puissance facturée vs consommée", fontsize=12, fontweight="bold")
        ax.set_ylabel("kW")
        ax.tick_params(axis="x", rotation=45)
        ax.legend(fontsize=8)
        plt.tight_layout()
        st.pyplot(fig)
        savefig(fig, "02_Puissance_facturee_vs_consommee")

# Tableau mensuel
with st.expander("📋 Tableau mensuel détaillé"):
    mon_display = mon.copy()
    mon_display.index = mon_display["Mois_label"]
    mon_display = mon_display.drop(columns=["Mois_label"])
    mon_display.columns = ["Énergie (MWh)", "P max (kW)", "P moy (kW)",
                            "FU moy (%)", "Heures couv.", "FU HQ (%)"]
    st.dataframe(mon_display.style.format({
        "Énergie (MWh)": "{:,.1f}",
        "P max (kW)":    "{:,.1f}",
        "P moy (kW)":    "{:,.1f}",
        "FU moy (%)":    "{:.1f}",
        "Heures couv.":  "{:,.0f}",
        "FU HQ (%)":     "{:.1f}",
    }), use_container_width=True)

# Consommation annuelle
ann = df["kWh"].resample("YE").sum() / 1000
if len(ann) > 0:
    with st.expander("📋 Consommation annuelle"):
        ann_df = ann.reset_index()
        ann_df.columns = ["Année", "Énergie (MWh)"]
        ann_df["Année"] = ann_df["Année"].dt.year
        ann_df["Énergie (MWh)"] = ann_df["Énergie (MWh)"].round(1)
        st.dataframe(ann_df, use_container_width=True)

# ─────────────────────────────────────────────────────────────────────────────
# SECTION 3 : ANALYSE SAISONNIÈRE
# ─────────────────────────────────────────────────────────────────────────────
st.divider()
st.subheader("🌿 Analyse saisonnière (Québec)")

ORDER_SAISON  = ["Hiver", "Printemps", "Été", "Automne"]
COL_SAISON    = {s: PALETTE[s.lower().replace("é","e")] for s in ORDER_SAISON}

sais = (
    df.groupby("Saison")
    .agg(
        E_MWh  =("kWh",  lambda x: x.sum() / 1000),
        P_max  =("kW",   "max"),
        P_moy  =("kW",   "mean"),
        FU_moy =("FU_%", "mean"),
        N      =("kW",   "count"),
    )
    .reindex(ORDER_SAISON)
    .dropna()
)

col_s1, col_s2 = st.columns(2)

with col_s1:
    fig, ax = plt.subplots(figsize=(8, 4))
    couleurs = [COL_SAISON.get(s, PALETTE["gris"]) for s in sais.index]
    bars = ax.bar(sais.index, sais["E_MWh"], color=couleurs, alpha=0.88, zorder=3)
    for bar, val in zip(bars, sais["E_MWh"]):
        ax.text(bar.get_x() + bar.get_width() / 2, bar.get_height() + 0.5,
                f"{val:,.0f}", ha="center", va="bottom", fontsize=8)
    ax.set_title("Consommation par saison (MWh)", fontsize=12, fontweight="bold")
    ax.set_ylabel("MWh")
    plt.tight_layout()
    st.pyplot(fig)
    savefig(fig, "03_Consommation_saisonniere")

with col_s2:
    fig, ax = plt.subplots(figsize=(8, 4))
    x = np.arange(len(sais))
    w = 0.35
    ax.bar(x - w/2, sais["P_max"], w, label="P max (kW)", color=PALETTE["rouge"],   alpha=0.85)
    ax.bar(x + w/2, sais["P_moy"], w, label="P moy (kW)", color=PALETTE["accent"],  alpha=0.85)
    ax.set_xticks(x); ax.set_xticklabels(sais.index)
    ax.set_title("Puissances max / moy par saison", fontsize=12, fontweight="bold")
    ax.set_ylabel("kW")
    ax.legend(fontsize=8)
    plt.tight_layout()
    st.pyplot(fig)
    savefig(fig, "04_Puissance_saisonniere")

# Tableau saison
sais_display = sais.copy()
sais_display.columns = ["Énergie (MWh)", "P max (kW)", "P moy (kW)", "FU moy (%)", "N points"]
st.dataframe(sais_display.style.format({
    "Énergie (MWh)": "{:,.1f}",
    "P max (kW)":    "{:,.1f}",
    "P moy (kW)":    "{:,.1f}",
    "FU moy (%)":    "{:.1f}",
    "N points":      "{:,.0f}",
}), use_container_width=True)

# Boxplot par saison
fig, ax = plt.subplots(figsize=(10, 4))
data_box = [df.loc[df["Saison"] == s, "kW"].dropna().values for s in ORDER_SAISON if s in df["Saison"].values]
labels_box = [s for s in ORDER_SAISON if s in df["Saison"].values]
bp = ax.boxplot(data_box, labels=labels_box, patch_artist=True, notch=False, showfliers=False)
for patch, lbl in zip(bp["boxes"], labels_box):
    patch.set_facecolor(COL_SAISON.get(lbl, PALETTE["gris"]))
    patch.set_alpha(0.75)
ax.axhline(palier, color=PALETTE["orange"], linestyle="--", linewidth=1.5,
           label=f"Palier {palier} kW")
ax.set_title("Distribution des puissances par saison (sans valeurs extrêmes)", fontsize=12, fontweight="bold")
ax.set_ylabel("kW")
ax.legend(fontsize=9)
plt.tight_layout()
st.pyplot(fig)
savefig(fig, "05_Boxplot_saisonnier")

# ─────────────────────────────────────────────────────────────────────────────
# SECTION 4 : PROFILS JOURNALIERS (jour type / semaine)
# ─────────────────────────────────────────────────────────────────────────────
st.divider()
st.subheader("🕐 Profil journalier moyen")

if not (10 <= step_min <= 20):
    st.warning(f"⚠️ Pas médian détecté = {step_min:.1f} min. Le profil horaire est optimal pour des données à ~15 min.")

# Profil global + par type de jour
prof = df.groupby(["Heure", "TypeJour"])["kW"].mean().unstack()
prof_saison = df.groupby(["Heure", "Saison"])["kW"].mean().unstack().reindex(ORDER_SAISON, axis=1)

col_p1, col_p2 = st.columns(2)

with col_p1:
    fig, ax = plt.subplots(figsize=(9, 4))
    if "Semaine" in prof.columns:
        ax.plot(prof.index, prof["Semaine"],  color=PALETTE["primaire"], lw=2.5, label="Semaine")
    if "Week-end" in prof.columns:
        ax.plot(prof.index, prof["Week-end"], color=PALETTE["orange"],   lw=2.5, label="Week-end", linestyle="--")
    ax.axhline(palier, color=PALETTE["rouge"], linestyle=":", lw=1.5, label=f"Palier {palier} kW")
    ax.set_title("Profil horaire moyen — Semaine vs Week-end", fontsize=12, fontweight="bold")
    ax.set_xlabel("Heure"); ax.set_ylabel("kW")
    ax.set_xticks(range(0, 24))
    ax.legend(fontsize=8)
    plt.tight_layout()
    st.pyplot(fig)
    savefig(fig, "06_Profil_semaine_vs_weekend")

with col_p2:
    fig, ax = plt.subplots(figsize=(9, 4))
    for saison_name in ORDER_SAISON:
        if saison_name in prof_saison.columns:
            ax.plot(prof_saison.index, prof_saison[saison_name],
                    color=COL_SAISON[saison_name], lw=2.2, label=saison_name)
    ax.axhline(palier, color=PALETTE["rouge"], linestyle=":", lw=1.5, label=f"Palier {palier} kW")
    ax.set_title("Profil horaire moyen par saison", fontsize=12, fontweight="bold")
    ax.set_xlabel("Heure"); ax.set_ylabel("kW")
    ax.set_xticks(range(0, 24))
    ax.legend(fontsize=8)
    plt.tight_layout()
    st.pyplot(fig)
    savefig(fig, "07_Profil_horaire_par_saison")

# Heatmap heure × jour de semaine
jours_labels = ["Lun", "Mar", "Mer", "Jeu", "Ven", "Sam", "Dim"]
heatmap_data = df.groupby(["Heure", "JourSemaine"])["kW"].mean().unstack()
heatmap_data.columns = [jours_labels[i] for i in heatmap_data.columns]

fig, ax = plt.subplots(figsize=(12, 5))
sns.heatmap(
    heatmap_data.T,
    cmap="YlOrRd",
    linewidths=0.3,
    linecolor="white",
    ax=ax,
    cbar_kws={"label": "kW moyen", "shrink": 0.8},
    fmt=".0f",
    annot=heatmap_data.T.shape[0] <= 7,
)
ax.set_title("Heatmap — Puissance moyenne par heure et jour de la semaine", fontsize=12, fontweight="bold")
ax.set_xlabel("Heure"); ax.set_ylabel("Jour")
plt.tight_layout()
st.pyplot(fig)
savefig(fig, "08_Heatmap_heure_jour")

# ─────────────────────────────────────────────────────────────────────────────
# SECTION 5 : FACTEUR DE CHARGE & FACTEUR D'UTILISATION
# ─────────────────────────────────────────────────────────────────────────────
st.divider()
st.subheader("⚙️ Facteur de charge & Facteur d'utilisation")

col_fu1, col_fu2 = st.columns(2)

with col_fu1:
    st.markdown("**Définitions**")
    st.markdown(f"""
- **Palier contractuel** : `{palier} kW` (calculé selon l'appel de pointe mesuré)
- **Facteur d'utilisation (FU)** = P_moy / Palier × 100 _(pondéré par le temps)_
- **Facteur de charge (FC)** = P_moy / P_max × 100 _(efficacité de l'utilisation)_
- Un FU élevé = bon usage du palier → moins de marge pour la réduction  
- Un FC faible = pointes courtes → potentiel d'écrêtage
""")

with col_fu2:
    fc_global = float(df["kW"].mean() / pmax * 100)
    d_fu = {
        "Indicateur": ["FU global (%)", "FC global (%)"],
        "Valeur":     [f"{fu:.1f}", f"{fc_global:.1f}"],
        "Interprétation": [
            "Bon (>70%) / Moyen (50-70%) / Faible (<50%)",
            "Bon (>70%) / Écrêtage possible (<60%)"
        ]
    }
    st.dataframe(pd.DataFrame(d_fu), use_container_width=True, hide_index=True)

# FU mensuel
if not mon.empty:
    fig, ax = plt.subplots(figsize=(10, 4))
    color_fu = [
        PALETTE["vert"]   if v >= 70 else
        PALETTE["orange"] if v >= 50 else
        PALETTE["rouge"]
        for v in mon["FU_moy_%"]
    ]
    bars = ax.bar(mon["Mois_label"], mon["FU_moy_%"], color=color_fu, alpha=0.85, zorder=3)
    ax.axhline(70, color=PALETTE["vert"],   linestyle="--", lw=1.5, label="Seuil bon (70%)")
    ax.axhline(50, color=PALETTE["orange"], linestyle="--", lw=1.5, label="Seuil moyen (50%)")
    for bar, val in zip(bars, mon["FU_moy_%"]):
        ax.text(bar.get_x() + bar.get_width() / 2, bar.get_height() + 0.3,
                f"{val:.0f}%", ha="center", va="bottom", fontsize=7)
    ax.set_title("Facteur d'utilisation mensuel (%)", fontsize=12, fontweight="bold")
    ax.set_ylabel("FU (%)")
    ax.set_ylim(0, 115)
    ax.tick_params(axis="x", rotation=45)
    ax.legend(fontsize=8)
    plt.tight_layout()
    st.pyplot(fig)
    savefig(fig, "09_FU_mensuel")

# ─────────────────────────────────────────────────────────────────────────────
# SECTION 6 : RÉPARTITION DES PUISSANCES
# ─────────────────────────────────────────────────────────────────────────────
st.divider()
st.subheader("📊 Répartition des puissances")

pmax_loc = float(df["kW"].max())
end_bin  = int(np.ceil(pmax_loc / bin_rep) * bin_rep) + bin_rep
bins_arr = np.arange(0, end_bin + bin_rep, bin_rep)
values   = df["kW"].dropna().values

fig, ax = plt.subplots(figsize=(12, 4))
n, bins_out, patches = ax.hist(
    values, bins=bins_arr,
    weights=np.ones_like(values) * 100.0 / len(values),
    rwidth=0.88, color=PALETTE["primaire"], alpha=0.85, zorder=3
)
# Colorier les barres au-dessus du palier en rouge
for patch, left_edge in zip(patches, bins_out[:-1]):
    if left_edge >= palier:
        patch.set_facecolor(PALETTE["rouge"])
        patch.set_alpha(0.8)

ax.axvline(palier, color=PALETTE["orange"], linestyle="--", lw=2,
           label=f"Palier {palier} kW")
ax.axvline(float(df["kW"].mean()), color=PALETTE["vert"], linestyle="-.", lw=1.8,
           label=f"Moyenne {df['kW'].mean():.0f} kW")
ax.set_title(f"Répartition des puissances (tranches de {bin_rep} kW)", fontsize=12, fontweight="bold")
ax.set_xlabel("Puissance (kW)"); ax.set_ylabel("% du temps")
ax.tick_params(axis="x", rotation=45)
ax.legend(fontsize=8)
plt.tight_layout()
st.pyplot(fig)
savefig(fig, "10_Repartition_puissance")

# Tableau répartition
df_tmp = df.copy()
df_tmp["Classe"] = pd.cut(df_tmp["kW"], bins=bins_arr, right=False, include_lowest=True)
rep = df_tmp["Classe"].value_counts(normalize=True).sort_index() * 100
df_rep = rep.reset_index()
df_rep.columns = [f"Classe ({bin_rep} kW)", "% du temps"]
df_rep["% du temps"] = df_rep["% du temps"].round(2)
with st.expander("📋 Tableau de répartition détaillé"):
    st.dataframe(df_rep, use_container_width=True, hide_index=True)

# ─────────────────────────────────────────────────────────────────────────────
# SECTION 7 : DÉTECTION D'ANOMALIES / PICS
# ─────────────────────────────────────────────────────────────────────────────
st.divider()
st.subheader("🚨 Détection des anomalies et pics de puissance")

mu    = float(df["kW"].mean())
sigma = float(df["kW"].std())
seuil = mu + seuil_anomalie * sigma

df_anom = df[df["kW"] >= seuil][["kW", "Saison", "TypeJour", "Heure"]].copy()
df_anom = df_anom.sort_values("kW", ascending=False).head(200)

st.markdown(
    f"Seuil actuel : **{seuil:,.1f} kW** "
    f"(= {mu:,.1f} + {seuil_anomalie} × {sigma:,.1f})  — "
    f"**{len(df_anom):,} événements** détectés"
)

col_a1, col_a2 = st.columns(2)

with col_a1:
    if not df_anom.empty:
        fig, ax = plt.subplots(figsize=(9, 4))
        ax.plot(df.index, df["kW"], color=PALETTE["gris"], lw=0.4, alpha=0.6, label="Série complète")
        ax.scatter(df_anom.index, df_anom["kW"],
                   color=PALETTE["rouge"], s=12, zorder=5, label="Pics détectés")
        ax.axhline(seuil,  color=PALETTE["rouge"],  linestyle="--", lw=1.5, label=f"Seuil {seuil:,.0f} kW")
        ax.axhline(palier, color=PALETTE["orange"],  linestyle=":",  lw=1.5, label=f"Palier {palier} kW")
        ax.set_title("Série temporelle avec anomalies détectées", fontsize=12, fontweight="bold")
        ax.set_ylabel("kW")
        ax.legend(fontsize=7)
        plt.tight_layout()
        st.pyplot(fig)
        savefig(fig, "11_Anomalies_serie_temporelle")
    else:
        st.info("Aucun pic détecté avec le seuil actuel.")

with col_a2:
    if not df_anom.empty:
        anom_par_saison = df_anom.groupby("Saison").size().reindex(ORDER_SAISON).fillna(0)
        fig, ax = plt.subplots(figsize=(7, 4))
        couleurs_s = [COL_SAISON.get(s, PALETTE["gris"]) for s in anom_par_saison.index]
        ax.bar(anom_par_saison.index, anom_par_saison.values, color=couleurs_s, alpha=0.85)
        ax.set_title("Pics par saison", fontsize=12, fontweight="bold")
        ax.set_ylabel("Nombre de pics")
        plt.tight_layout()
        st.pyplot(fig)
        savefig(fig, "12_Anomalies_par_saison")

if not df_anom.empty:
    with st.expander("📋 Top 50 pics — détail"):
        st.dataframe(
            df_anom.head(50).reset_index().rename(columns={"ts": "Timestamp", "kW": "Puissance (kW)"}),
            use_container_width=True,
            hide_index=True
        )

# ─────────────────────────────────────────────────────────────────────────────
# SECTION 8 : EXPORT EXCEL (stylistique + données brutes)
# ─────────────────────────────────────────────────────────────────────────────
st.divider()
st.subheader("📤 Export Excel")

def build_excel() -> bytes:
    wb = Workbook()
    wb.remove(wb.active)

    sous_titre = f"Période : {periode}  |  Pas médian : {step_min:.1f} min  |  Palier : {palier} kW"

    # ── Onglet 1 : KPI Rapport (stylisé)
    ws_kpi = wb.create_sheet("Rapport KPI")
    kpi_data = pd.DataFrame([{
        "Indicateur":              "Appel de pointe mesuré (kW)",
        "Valeur":                  round(pmax, 1),
        "Commentaire":             f"Max sur la période {periode}"
    }, {
        "Indicateur":              "Palier contractuel (kW)",
        "Valeur":                  palier,
        "Commentaire":             "Calculé selon appel de pointe"
    }, {
        "Indicateur":              "Énergie totale (MWh)",
        "Valeur":                  round(e_kwh / 1000, 1),
        "Commentaire":             ""
    }, {
        "Indicateur":              "Heures couvertes (h)",
        "Valeur":                  round(h_total, 0),
        "Commentaire":             ""
    }, {
        "Indicateur":              "Facteur d'utilisation FU (%)",
        "Valeur":                  round(fu, 1),
        "Commentaire":             "FU pondéré par le temps"
    }, {
        "Indicateur":              "Facteur de charge FC (%)",
        "Valeur":                  round(float(df["kW"].mean()) / pmax * 100, 1),
        "Commentaire":             "P_moy / P_max × 100"
    }, {
        "Indicateur":              "Pas médian (min)",
        "Valeur":                  round(step_min, 1),
        "Commentaire":             "Médiane des intervalles de temps"
    }])
    xl_write_df(ws_kpi, "RAPPORT — INDICATEURS CLÉS", sous_titre, "⚡  KPI GLOBAUX", kpi_data)

    # ── Onglet 2 : Mensuel (stylisé)
    if not mon.empty:
        ws_mon = wb.create_sheet("Mensuel")
        mon_xl = mon.copy()
        mon_xl.index = mon_xl["Mois_label"]
        mon_xl = mon_xl.drop(columns=["Mois_label"]).reset_index()
        mon_xl.columns = ["Mois", "Énergie MWh", "P max kW", "P moy kW",
                          "FU moy %", "Heures couv", "FU HQ %"]
        mon_xl = mon_xl.round(2)
        xl_write_df(ws_mon, "ANALYSE MENSUELLE — HYDRO-QUÉBEC", sous_titre,
                    "📅  CONSOMMATION ET PUISSANCES PAR MOIS", mon_xl)

    # ── Onglet 3 : Saisonnier (stylisé)
    ws_sais = wb.create_sheet("Saisonnier")
    sais_xl = sais.copy().reset_index()
    sais_xl.columns = ["Saison", "Énergie MWh", "P max kW", "P moy kW", "FU moy %", "N points"]
    sais_xl = sais_xl.round(2)
    xl_write_df(ws_sais, "ANALYSE SAISONNIÈRE — HYDRO-QUÉBEC", sous_titre,
                "🌿  HIVER / PRINTEMPS / ÉTÉ / AUTOMNE", sais_xl)

    # ── Onglet 4 : Répartition puissances (stylisé)
    ws_rep = wb.create_sheet("Répartition")
    xl_write_df(ws_rep, f"RÉPARTITION DES PUISSANCES ({bin_rep} kW)", sous_titre,
                "📊  DISTRIBUTION DES APPELS DE PUISSANCE", df_rep.round(2))

    # ── Onglet 5 : Anomalies (stylisé)
    if not df_anom.empty:
        ws_anom = wb.create_sheet("Anomalies")
        anom_xl = df_anom.head(200).reset_index()
        anom_xl.columns = ["Timestamp", "Puissance (kW)", "Saison", "TypeJour", "Heure"]
        anom_xl["Puissance (kW)"] = anom_xl["Puissance (kW)"].round(1)
        xl_write_df(ws_anom, "PICS ET ANOMALIES DÉTECTÉS", sous_titre,
                    f"🚨  SEUIL : {seuil:,.1f} kW  (moy + {seuil_anomalie}σ)", anom_xl)

    # ── Onglet 6 : Profil horaire (stylisé)
    ws_prof = wb.create_sheet("Profil horaire")
    prof_xl = df.groupby(["Heure", "TypeJour", "Saison"])["kW"].mean().reset_index()
    prof_xl["kW"] = prof_xl["kW"].round(1)
    prof_xl.columns = ["Heure", "Type jour", "Saison", "kW moyen"]
    xl_write_df(ws_prof, "PROFIL HORAIRE MOYEN", sous_titre,
                "🕐  PUISSANCE MOYENNE PAR HEURE / TYPE DE JOUR / SAISON", prof_xl)

    # ── Onglet 7 : Données brutes (non stylisé — données complètes)
    ws_raw = wb.create_sheet("Données brutes")
    df_raw_xl = df.reset_index()[["ts", "kW", "kWh", "delta_h",
                                   "FU_%", "Saison", "TypeJour", "Heure", "_src"]].copy()
    df_raw_xl.columns = ["Timestamp", "Puissance (kW)", "Énergie (kWh)", "Intervalle (h)",
                          "FU (%)", "Saison", "Type jour", "Heure", "Source fichier"]
    df_raw_xl = df_raw_xl.round(4)
    df_raw_xl.to_excel(ws_raw, index=False, startrow=0)
    # En-tête stylisé minimal
    ws_raw_style = wb["Données brutes"]
    for ci, h in enumerate(df_raw_xl.columns, 1):
        c = ws_raw_style.cell(row=1, column=ci)
        c.font = Font(name=XL_FONT, bold=True, size=10, color=XL_WHITE)
        c.fill = PatternFill("solid", fgColor=XL_HEADER_BG)
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = _med_bottom()
        ws_raw_style.column_dimensions[get_column_letter(ci)].width = max(len(str(h)) + 4, 14)
    ws_raw_style.freeze_panes = ws_raw_style["A2"]

    # ── Onglet 8 : Graphiques (images)
    ws_graphs = wb.create_sheet("Graphiques")
    xl_titre(ws_graphs, 1, "GRAPHIQUES — ANALYSE HYDRO-QUÉBEC", 4)
    xl_sous_titre(ws_graphs, 2, sous_titre, 4)
    ws_graphs.row_dimensions[1].height = 22
    ws_graphs.row_dimensions[2].height = 16

    graph_row = 4
    for key, buf in fig_buffers.items():
        ws_graphs.cell(row=graph_row, column=1, value=key.replace("_", " "))
        ws_graphs.cell(row=graph_row, column=1).font = Font(name=XL_FONT, bold=True, size=10, color=XL_SECTION_BG)
        graph_row += 1

        img = XLImage(buf)
        img.width  = 750
        img.height = 350
        img.anchor = f"A{graph_row}"
        ws_graphs.add_image(img)
        graph_row += 24

    buf_out = io.BytesIO()
    wb.save(buf_out)
    buf_out.seek(0)
    return buf_out.read()


excel_bytes = build_excel()

st.download_button(
    label="📥 Télécharger la synthèse Excel complète",
    data=excel_bytes,
    file_name="Analyse_HydroQuebec.xlsx",
    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
)

st.caption(
    "Export contient : KPI · Mensuel · Saisonnier · Répartition puissances · "
    "Anomalies · Profil horaire · Données brutes · Graphiques"
)
